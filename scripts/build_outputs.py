#!/usr/bin/env python3
"""
Validate a Claude-authored dataset.json against its declared schema and emit the
final deliverables: dataset.csv, dataset.json (normalized/re-serialized), dataset.xlsx
(if openpyxl is available), and validation_notes.json (machine-readable list of
schema violations for the agent to fold into the extraction report).

This script does NOT infer the schema or extract field values -- that reasoning step
is done by the Claude agent, which is much better suited to judgment calls like "what
is the unit of analysis" or "does this value contradict that one". This script's job
is the boring, deterministic, easy-to-get-wrong-by-hand part: consistent CSV/JSON/XLSX
serialization and mechanical validation.

Expected input dataset.json shape:
{
  "schema": [
    {"name": "sample_id", "description": "...", "type": "string",
     "required": true, "examples": ["S-01", "S-02"]}
  ],
  "records": [
    {
      "_record_id": "rec-001",
      "_duplicate_of": null,
      "_duplicate_confidence": null,
      "field_name": {
        "value": <normalized value or null>,
        "value_raw": <original as-seen string, or null if identical/not applicable>,
        "status": "stated" | "normalized" | "inferred" | "missing",
        "confidence": 0.0-1.0 or null,
        "provenance": [{"file": "...", "location": "page 2", "quote": "..."}],
        "conflict": null or {"description": "...", "values": [...]}
      },
      ...
    }
  ]
}

Usage:
    python3 build_outputs.py <dataset_json> <output_dir>
"""
import sys
import os
import json
import csv


def load_dataset(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def validate(dataset):
    """Returns a list of validation issues (dicts). Never raises."""
    issues = []
    schema = dataset.get('schema', [])
    records = dataset.get('records', [])
    field_names = [f['name'] for f in schema]

    if not schema:
        issues.append({"level": "error", "message": "dataset.json has no 'schema' array"})
    if not records:
        issues.append({"level": "warning", "message": "dataset.json has no 'records'"})

    seen_ids = set()
    for i, rec in enumerate(records):
        rid = rec.get('_record_id', f'<index {i}>')
        if rid in seen_ids:
            issues.append({"level": "error", "record": rid, "message": "duplicate _record_id"})
        seen_ids.add(rid)

        for f in schema:
            name = f['name']
            required = f.get('required', False)
            cell = rec.get(name)
            if cell is None:
                if required:
                    issues.append({"level": "error", "record": rid,
                                    "message": f"missing required field '{name}'"})
                continue
            if not isinstance(cell, dict):
                issues.append({"level": "error", "record": rid,
                                "message": f"field '{name}' is not an object with value/status/provenance"})
                continue
            status = cell.get('status')
            if status not in ('stated', 'normalized', 'inferred', 'missing'):
                issues.append({"level": "warning", "record": rid,
                                "message": f"field '{name}' has unexpected status '{status}'"})
            if status == 'missing' and cell.get('value') is not None:
                issues.append({"level": "warning", "record": rid,
                                "message": f"field '{name}' marked missing but has a non-null value"})
            if required and status == 'missing':
                # A genuinely missing/conflicted value in a required field is a real
                # finding to surface in the report, not a reason to block output
                # generation -- the whole point of this workflow is to make gaps
                # visible rather than papering over them with a fabricated value.
                issues.append({"level": "warning", "record": rid,
                                "message": f"required field '{name}' is missing"})
            if cell.get('confidence') is not None:
                try:
                    c = float(cell['confidence'])
                    if not (0.0 <= c <= 1.0):
                        issues.append({"level": "warning", "record": rid,
                                        "message": f"field '{name}' confidence {c} outside [0,1]"})
                except (TypeError, ValueError):
                    issues.append({"level": "warning", "record": rid,
                                    "message": f"field '{name}' confidence is not numeric"})
            if status in ('stated', 'normalized') and not cell.get('provenance'):
                issues.append({"level": "warning", "record": rid,
                                "message": f"field '{name}' has status '{status}' but no provenance"})

        for key in rec.keys():
            if key.startswith('_'):
                continue
            if key not in field_names:
                issues.append({"level": "warning", "record": rid,
                                "message": f"field '{key}' present on record but not declared in schema"})

    return issues


def flatten_value(cell):
    if cell is None:
        return ''
    if not isinstance(cell, dict):
        return cell
    v = cell.get('value')
    if v is None:
        return ''
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def write_csv(dataset, out_path):
    schema = dataset.get('schema', [])
    records = dataset.get('records', [])
    field_names = [f['name'] for f in schema]
    header = ['_record_id'] + field_names + ['_duplicate_of', '_duplicate_confidence']
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(header)
        for rec in records:
            row = [rec.get('_record_id', '')]
            for name in field_names:
                row.append(flatten_value(rec.get(name)))
            row.append(rec.get('_duplicate_of') or '')
            row.append(rec.get('_duplicate_confidence') or '')
            w.writerow(row)


def write_json(dataset, out_path):
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(dataset, f, indent=2, ensure_ascii=False)


def write_xlsx(dataset, out_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False, "openpyxl not installed; skipped dataset.xlsx (dataset.csv/json are authoritative)"

    schema = dataset.get('schema', [])
    records = dataset.get('records', [])
    field_names = [f['name'] for f in schema]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "dataset"
    header = ['_record_id'] + field_names + ['_duplicate_of', '_duplicate_confidence']
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    low_conf_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    conflict_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for rec in records:
        row = [rec.get('_record_id', '')]
        flags = []
        for name in field_names:
            cell = rec.get(name)
            row.append(flatten_value(cell))
            flag = None
            if isinstance(cell, dict):
                if cell.get('conflict'):
                    flag = 'conflict'
                elif isinstance(cell.get('confidence'), (int, float)) and cell['confidence'] < 0.6:
                    flag = 'low_confidence'
            flags.append(flag)
        row.append(rec.get('_duplicate_of') or '')
        row.append(rec.get('_duplicate_confidence') or '')
        ws.append(row)
        r = ws.max_row
        for col_idx, flag in enumerate(flags, start=2):
            if flag == 'conflict':
                ws.cell(row=r, column=col_idx).fill = conflict_fill
            elif flag == 'low_confidence':
                ws.cell(row=r, column=col_idx).fill = low_conf_fill

    # schema sheet
    ws2 = wb.create_sheet("schema")
    ws2.append(["name", "description", "type", "required", "examples"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for f in schema:
        ws2.append([
            f.get('name', ''), f.get('description', ''), f.get('type', ''),
            f.get('required', False),
            ', '.join(str(x) for x in f.get('examples', [])) if f.get('examples') else ''
        ])

    for sheet in (ws, ws2):
        for col in sheet.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 60)

    wb.save(out_path)
    return True, None


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_outputs.py <dataset_json> <output_dir>", file=sys.stderr)
        sys.exit(2)
    dataset_path, out_dir = sys.argv[1], sys.argv[2]
    os.makedirs(out_dir, exist_ok=True)

    dataset = load_dataset(dataset_path)
    issues = validate(dataset)

    write_csv(dataset, os.path.join(out_dir, 'dataset.csv'))
    write_json(dataset, os.path.join(out_dir, 'dataset.json'))
    xlsx_ok, xlsx_note = write_xlsx(dataset, os.path.join(out_dir, 'dataset.xlsx'))

    notes = {
        "validation_issues": issues,
        "xlsx_generated": xlsx_ok,
        "xlsx_note": xlsx_note,
        "record_count": len(dataset.get('records', [])),
        "field_count": len(dataset.get('schema', [])),
    }
    with open(os.path.join(out_dir, 'validation_notes.json'), 'w', encoding='utf-8') as f:
        json.dump(notes, f, indent=2, ensure_ascii=False)

    print(f"Wrote dataset.csv, dataset.json to {out_dir}")
    print(f"dataset.xlsx: {'written' if xlsx_ok else 'skipped (' + str(xlsx_note) + ')'}")
    errors = [i for i in issues if i.get('level') == 'error']
    warnings = [i for i in issues if i.get('level') == 'warning']
    print(f"Validation: {len(errors)} error(s), {len(warnings)} warning(s). See validation_notes.json.")
    if errors:
        sys.exit(1)


if __name__ == '__main__':
    main()
